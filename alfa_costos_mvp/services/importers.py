from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from alfa_costos_mvp.models import ImportFile, ImportedRow, SourceKind
from alfa_costos_mvp.services.normalizers import parse_decimal


class ImportErrorDetail(Exception):
    pass


class BaseImporter:
    def read(self, import_file: ImportFile) -> Iterable[ImportedRow]:
        raise NotImplementedError


class StructuredFileImporter(BaseImporter):
    """
    Punto de entrada para Excel, CSV y TXT.
    En el MVP real deberia:
    - detectar hojas / delimitadores
    - sugerir columnas
    - devolver filas normalizadas
    """

    CODE_CANDIDATES = ("codigo", "cod", "sku", "codigo proveedor", "idarticuloproveedor")
    DESCRIPTION_CANDIDATES = ("descripcion", "descripción", "detalle", "producto", "articulo")
    PRICE_CANDIDATES = (
        "precio s/iva",
        "precio sin iva",
        "precio costo",
        "costo",
        "precio",
        "importe",
        "precio c/iva",
        "precio con iva",
    )

    def read(self, import_file: ImportFile) -> Iterable[ImportedRow]:
        if import_file.source_kind not in {SourceKind.EXCEL, SourceKind.CSV, SourceKind.TXT}:
            raise ImportErrorDetail("Tipo de archivo no soportado por lectura estructurada.")
        path = import_file.path
        if path.suffix.lower() == ".xlsx":
            return self._read_xlsx(import_file)
        if path.suffix.lower() == ".xls":
            raise ImportErrorDetail(
                "El soporte robusto para .xls todavia no esta implementado en el MVP. "
                "Conviene convertir a .xlsx o usar CSV/TXT."
            )
        return self._read_delimited(import_file)

    def _read_xlsx(self, import_file: ImportFile) -> list[ImportedRow]:
        workbook = load_workbook(import_file.path, read_only=True, data_only=True)
        rows: list[ImportedRow] = []
        preferred_sheet = (import_file.sheet_name or "").strip().upper()

        worksheets = list(workbook.worksheets)
        if preferred_sheet:
            ordered = [ws for ws in worksheets if ws.title.strip().upper() == preferred_sheet]
            ordered += [ws for ws in worksheets if ws.title.strip().upper() != preferred_sheet]
        else:
            ordered = worksheets

        for ws in ordered:
            parsed = self._parse_matrix(
                matrix=list(ws.iter_rows(values_only=True)),
                sheet_name=ws.title,
            )
            if parsed:
                rows.extend(parsed)
                break
        return rows

    def _read_delimited(self, import_file: ImportFile) -> list[ImportedRow]:
        encodings = ("utf-8-sig", "latin-1", "cp1252")
        last_error: Optional[Exception] = None
        for encoding in encodings:
            try:
                with import_file.path.open("r", encoding=encoding, newline="") as handle:
                    sample = handle.read(4096)
                    handle.seek(0)
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
                    reader = csv.reader(handle, dialect)
                    matrix = [row for row in reader]
                return self._parse_matrix(matrix=matrix, sheet_name=None)
            except Exception as exc:
                last_error = exc
        raise ImportErrorDetail(f"No se pudo leer archivo delimitado: {last_error}")

    def _parse_matrix(self, matrix: list[tuple | list], sheet_name: Optional[str]) -> list[ImportedRow]:
        header_idx, header_map = self._detect_header(matrix)
        if header_idx is None or not header_map:
            return []

        code_idx = header_map.get("provider_code")
        description_idx = header_map.get("description")
        price_idx = header_map.get("cost_price")
        if description_idx is None or price_idx is None:
            return []

        imported_rows: list[ImportedRow] = []
        for offset, raw_row in enumerate(matrix[header_idx + 1 :], start=header_idx + 2):
            values = ["" if value is None else str(value).strip() for value in raw_row]
            if not any(values):
                continue

            description = values[description_idx].strip() if description_idx < len(values) else ""
            if not description:
                continue

            raw_price = values[price_idx].strip() if price_idx < len(values) else ""
            cost_price = parse_decimal(raw_price)
            if cost_price is None:
                continue

            provider_code = values[code_idx].strip() if code_idx is not None and code_idx < len(values) else ""
            imported_rows.append(
                ImportedRow(
                    row_number=offset,
                    provider_code=provider_code,
                    description=description,
                    cost_price=cost_price,
                    raw_code=provider_code,
                    raw_description=description,
                    raw_price=raw_price,
                    selected_price_column="cost_price",
                    source_sheet=sheet_name,
                    raw_values={
                        "provider_code": provider_code,
                        "description": description,
                        "cost_price": raw_price,
                    },
                )
            )
        return imported_rows

    def _detect_header(self, matrix: list[tuple | list]) -> tuple[Optional[int], dict[str, int]]:
        for row_idx, row in enumerate(matrix[:20]):
            normalized = [self._normalize_header_cell(value) for value in row]
            if not any(normalized):
                continue

            header_map: dict[str, int] = {}
            for col_idx, value in enumerate(normalized):
                if not value:
                    continue
                if "provider_code" not in header_map and any(token in value for token in self.CODE_CANDIDATES):
                    header_map["provider_code"] = col_idx
                if "description" not in header_map and any(token in value for token in self.DESCRIPTION_CANDIDATES):
                    header_map["description"] = col_idx
                if "cost_price" not in header_map and any(token in value for token in self.PRICE_CANDIDATES):
                    header_map["cost_price"] = col_idx

            if "description" in header_map and "cost_price" in header_map:
                return row_idx, header_map
        return None, {}

    @staticmethod
    def _normalize_header_cell(value: object) -> str:
        text = "" if value is None else str(value).strip().lower()
        return " ".join(text.split())


class DocumentAIImporter(BaseImporter):
    """
    Adaptador para PDF e imagen.
    Debe reutilizar piezas del lector actual, no copiar toda su logica.
    """

    def read(self, import_file: ImportFile) -> Iterable[ImportedRow]:
        if import_file.source_kind not in {SourceKind.PDF, SourceKind.IMAGE}:
            raise ImportErrorDetail("Tipo de archivo no soportado por lectura documental.")
        return []


def detect_source_kind(path: Path) -> SourceKind:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xls"}:
        return SourceKind.EXCEL
    if ext == ".csv":
        return SourceKind.CSV
    if ext == ".txt":
        return SourceKind.TXT
    if ext == ".pdf":
        return SourceKind.PDF
    if ext in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}:
        return SourceKind.IMAGE
    raise ImportErrorDetail(f"Extension no soportada: {ext}")
